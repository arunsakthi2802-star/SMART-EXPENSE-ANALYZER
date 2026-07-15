from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ReportService:
    @staticmethod
    def generate_excel_report(transactions: List[Dict[str, Any]], user_name: str) -> BytesIO:
        """
        Generates a professionally styled Excel spreadsheet containing transactions and summary metrics.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions Ledger"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Colors & Fills
        primary_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark Blue
        accent_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")    # Gray accent
        income_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")    # Mint green
        expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Soft red
        
        # Fonts
        title_font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Segoe UI", size=11, bold=True)
        regular_font = Font(name="Segoe UI", size=11)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='9CA3AF'),
            bottom=Side(style='double', color='1E3A8A')
        )

        # Title Block
        ws.merge_cells("A1:G1")
        ws["A1"] = "SMART EXPENSE ANALYZER - TRANSACTIONS LEDGER"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws["A2"] = f"Generated for: {user_name}"
        ws["A2"].font = bold_font
        ws["A3"] = f"Date Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"].font = regular_font
        
        # Leave a blank row
        ws.append([])
        
        # Summary metrics
        total_income = sum(t['amount'] for t in transactions if t['type'] == 'income')
        total_expense = sum(t['amount'] for t in transactions if t['type'] == 'expense')
        net_savings = total_income - total_expense
        
        ws["A5"] = "Total Income"
        ws["B5"] = total_income
        ws["A5"].font = bold_font
        ws["B5"].font = bold_font
        ws["B5"].number_format = "$#,##0.00"
        
        ws["D5"] = "Total Expenses"
        ws["E5"] = total_expense
        ws["D5"].font = bold_font
        ws["E5"].font = bold_font
        ws["E5"].number_format = "$#,##0.00"
        
        ws["G5"] = "Net Savings"
        ws["H5"] = net_savings
        ws["G5"].font = bold_font
        ws["H5"].font = bold_font
        ws["H5"].number_format = "$#,##0.00"
        
        ws.append([])
        ws.append([])
        
        # Table Headers
        headers = ["Date", "Type", "Category", "Merchant", "Description", "Payment Method", "Amount"]
        ws.append(headers)
        
        header_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = primary_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        # Add Data rows
        for t in sorted(transactions, key=lambda x: x['date']):
            # Convert date representation to string if it is a datetime object
            date_str = t['date'].strftime('%Y-%m-%d') if isinstance(t['date'], datetime) else str(t['date'])[:10]
            row_data = [
                date_str,
                t['type'].upper(),
                t.get('category_name', 'N/A'),
                t.get('merchant_name', ''),
                t.get('description', ''),
                t.get('payment_method', 'Cash'),
                t['amount']
            ]
            ws.append(row_data)
            current_row = ws.max_row
            
            # Format rows
            type_cell = ws.cell(row=current_row, column=2)
            type_cell.fill = income_fill if t['type'] == 'income' else expense_fill
            type_cell.alignment = Alignment(horizontal="center")
            
            amount_cell = ws.cell(row=current_row, column=7)
            amount_cell.number_format = "$#,##0.00"
            amount_cell.alignment = Alignment(horizontal="right")
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border

        # Adjust column dimensions dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format == "$#,##0.00":
                    val = f"${float(cell.value or 0):,.2f}"
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream

    @staticmethod
    def generate_pdf_report(
        transactions: List[Dict[str, Any]], 
        user_name: str, 
        period_label: str = "All Time",
        ai_insights: Dict[str, Any] = None
    ) -> BytesIO:
        """
        Generates a corporate-grade financial statement PDF using ReportLab flowables.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Custom Typography Styles
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#1E3A8A"),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            name="SubTitleStyle",
            parent=styles["Normal"],
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=20
        )
        h2_style = ParagraphStyle(
            name="H2Style",
            parent=styles["Heading2"],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#111827"),
            spaceBefore=12,
            spaceAfter=8
        )
        body_style = ParagraphStyle(
            name="BodyStyle",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#1F2937")
        )
        body_bold = ParagraphStyle(
            name="BodyBold",
            parent=body_style,
            fontName="Helvetica-Bold"
        )
        header_table_style = ParagraphStyle(
            name="HeaderTableStyle",
            parent=body_style,
            textColor=colors.white,
            fontName="Helvetica-Bold"
        )

        elements = []
        
        # 1. Header block
        elements.append(Paragraph("SMART EXPENSE ANALYZER", title_style))
        elements.append(Paragraph(f"Financial Statement | Prepared for: {user_name} | Scope: {period_label}", subtitle_style))
        elements.append(Spacer(1, 10))

        # 2. Key Metrics Summary Grid
        total_income = sum(t['amount'] for t in transactions if t['type'] == 'income')
        total_expense = sum(t['amount'] for t in transactions if t['type'] == 'expense')
        net_savings = total_income - total_expense
        savings_ratio = (net_savings / total_income * 100) if total_income > 0 else 0
        
        summary_data = [
            [
                Paragraph("TOTAL INCOME", body_bold),
                Paragraph("TOTAL EXPENSES", body_bold),
                Paragraph("NET SAVINGS", body_bold),
                Paragraph("SAVINGS RATE", body_bold)
            ],
            [
                Paragraph(f"${total_income:,.2f}", ParagraphStyle('In', parent=body_style, textColor=colors.HexColor("#10B981"), fontSize=14, leading=16)),
                Paragraph(f"${total_expense:,.2f}", ParagraphStyle('Ex', parent=body_style, textColor=colors.HexColor("#EF4444"), fontSize=14, leading=16)),
                Paragraph(f"${net_savings:,.2f}", ParagraphStyle('Net', parent=body_style, textColor=colors.HexColor("#1E3A8A"), fontSize=14, leading=16)),
                Paragraph(f"{savings_ratio:.1f}%", ParagraphStyle('Sr', parent=body_style, textColor=colors.HexColor("#3B82F6"), fontSize=14, leading=16))
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[130, 130, 130, 130])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F9FAFB")),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#E5E7EB")),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#F3F4F6"))
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 15))

        # AI Financial Analysis & Insights
        if ai_insights:
            elements.append(Paragraph("AI Financial Analysis & Insights", h2_style))
            
            score = ai_insights.get("score", 0)
            breakdown = ai_insights.get("breakdown", {})
            
            # Format custom summary box
            ai_score_text = (
                f"<b>Overall Financial Score:</b> {score}/100  |  "
                f"<b>Savings Score:</b> {breakdown.get('savings_score', 0)}/40  |  "
                f"<b>Budget Compliance:</b> {breakdown.get('budget_compliance_score', 0)}/30  |  "
                f"<b>Priority Score:</b> {breakdown.get('priority_score', 0)}/30"
            )
            elements.append(Paragraph(ai_score_text, body_bold))
            elements.append(Spacer(1, 6))
            
            recs = ai_insights.get("recommendations", [])
            if recs:
                elements.append(Paragraph("<b>AI Recommendations & Actions:</b>", body_bold))
                for r in recs:
                    elements.append(Paragraph(f"• {r}", ParagraphStyle('RecStyle', parent=body_style, leftIndent=12, spaceAfter=3)))
            else:
                elements.append(Paragraph("No current recommendations generated.", body_style))
                
            elements.append(Spacer(1, 15))

        # Visual Analytics Hub (Category Breakdown)
        total_expense = sum(t['amount'] for t in transactions if t['type'] == 'expense')
        if total_expense > 0:
            elements.append(Paragraph("Visual Analytics Hub - Category Breakdown", h2_style))
            cat_map = {}
            for t in transactions:
                if t['type'] == 'expense':
                    cat = t.get('category_name', 'Other')
                    cat_map[cat] = cat_map.get(cat, 0.0) + t['amount']
                    
            analytics_data = [[
                Paragraph("Category", header_table_style),
                Paragraph("Amount Spent", header_table_style),
                Paragraph("Percentage of Outflows", header_table_style)
            ]]
            
            for cat, amt in sorted(cat_map.items(), key=lambda x: x[1], reverse=True):
                pct = (amt / total_expense * 100)
                analytics_data.append([
                    Paragraph(cat, body_style),
                    Paragraph(f"${amt:,.2f}", body_style),
                    Paragraph(f"{pct:.1f}%", body_style)
                ])
                
            analytics_table = Table(analytics_data, colWidths=[180, 180, 180])
            analytics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3B82F6")),  # Blue header
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ]))
            elements.append(analytics_table)
            elements.append(Spacer(1, 15))

        # 3. Transaction Details Ledger
        elements.append(Paragraph("Transaction History", h2_style))
        
        table_data = [[
            Paragraph("Date", header_table_style),
            Paragraph("Type", header_table_style),
            Paragraph("Category", header_table_style),
            Paragraph("Merchant", header_table_style),
            Paragraph("Method", header_table_style),
            Paragraph("Amount", header_table_style)
        ]]
        
        for t in sorted(transactions, key=lambda x: x['date'], reverse=True):
            date_str = t['date'].strftime('%Y-%m-%d') if isinstance(t['date'], datetime) else str(t['date'])[:10]
            type_str = t['type'].upper()
            
            amount_style = ParagraphStyle(
                'amt',
                parent=body_style,
                textColor=colors.HexColor("#10B981") if t['type'] == 'income' else colors.HexColor("#EF4444")
            )
            
            table_data.append([
                Paragraph(date_str, body_style),
                Paragraph(type_str, body_style),
                Paragraph(t.get('category_name', 'N/A'), body_style),
                Paragraph(t.get('merchant_name', 'N/A') or '-', body_style),
                Paragraph(t.get('payment_method', 'Cash'), body_style),
                Paragraph(f"${t['amount']:,.2f}", amount_style)
            ])
            
        # Total transaction width available: 540 (612 - 72 margins)
        ledger_table = Table(table_data, colWidths=[75, 55, 110, 130, 80, 90])
        ledger_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ]))
        
        elements.append(ledger_table)
        
        # Build Document
        doc.build(elements)
        buffer.seek(0)
        return buffer
